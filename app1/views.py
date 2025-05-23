import os
import random
import threading
import pandas as pd

from io import BytesIO
from openpyxl import load_workbook

from django.shortcuts import get_object_or_404
from django.conf import settings
from django.http import FileResponse, HttpResponse
from django.http import JsonResponse
from django.core.mail import EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.http import require_GET
from django_filters.rest_framework import DjangoFilterBackend, FilterSet, CharFilter

from threading import Thread

from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework import status
from rest_framework import viewsets
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework import generics
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView

from datetime import timedelta

# All import Models files 
from .models import (
    CustomUserModel,
    CustomGroupModel,
    PendingUser,
    ExpiringToken,
    Product,
    Category,
    ProductCategory,
    Order,
    OrderItem,
    Payment,
    Store,
)

# All Serializers import files
from .serializers import (
    RegisterSerializer,
    LoginSerializer,
    # AddUsrToGrpSerializer,
    OTPVerifySerializer,
    UserSerializer,
    GroupWithUsersSerializer,
    GroupCreateSerializers,
    GroupSerializer,
    ProductSerializer,
    CategorySerializer,
    ProductCategorySerializer,
    OrderSerializer,
    PaymentSerializer,
    ProductExcelSerializers,
    CategoryExcelSerializer,
    StoreSerializer,
    StoreCreateUpdateSerializer,
    StoreToggleSerializer,
)
# from .signals import order_confirmed, create_loyalty_point
# from django.core.mail import EmailMultiAlternatives
# from django.core.mail import send_mail
# from rest_framework.authtoken.models import Token
# from django.db import transaction
# from rest_framework.viewsets import ModelViewSet


# simple Registration and Login with API
# Handle user registration and send OTP to email
@api_view(["POST"])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        email = serializer.validated_data["email"]
        username = serializer.validated_data["username"]

        # Check if user is already waiting for OTP verification
        if PendingUser.objects.filter(username=username).exists():
            return Response({"error": "Email already registered for OTP."}, status=400)

        # Generate 4-digit OTP and save the temporary user
        otp = str(random.randint(1000, 9999))
        serializer.save(otp=otp)

        # Send OTP email to the user
        subject = "Please verify your email with OTP"
        # message = f'Hi {username},\nYour OTP is: {otp}'

        from_email = "abaranwal.it@gmail.com"
        to_email = [email]

        context = {"otp": otp, "username": username}
        html_content = render_to_string("app1/emails/otp.html", context)
        text_content = "OTP  here!"

        # Send the email for old option
        # email = EmailMultiAlternatives(subject, text_content, from_email, to_email)
        # email.attach_alternative(html_content, "text/html")
        # email.send()

        def send_otp_email():
            try:
                email_message = EmailMultiAlternatives(
                    subject, text_content, from_email, to_email
                )
                email_message.attach_alternative(html_content, "text/html")
                email_message.send(fail_silently = False)
            except Exception as e:
                print("Error sending email:", e)

        threading.Thread(target = send_otp_email).start()

        return Response({"message" : "OTP sent to email"}, status = status.HTTP_200_OK)
    return Response(serializer.errors, status = status.HTTP_400_BAD_REQUEST)


# Verify OTP and create the actual user account
@api_view(["POST"])
def verify_otp(request):
    serializer = OTPVerifySerializer(data=request.data)
    if serializer.is_valid():
        email = serializer.validated_data["email"]
        otp = serializer.validated_data["otp"]

        # Fetch pending user based on email
        try:
            pending_user = PendingUser.objects.get(email=email)
        except PendingUser.DoesNotExist:
            return Response({"error": "No pending user found."}, status=404)

        # Compare submitted OTP with saved one
        if pending_user.otp != otp:
            return Response({"error": "Incorrect OTP"}, status=400)

        # Create actual user from pending user data
        user = CustomUserModel.objects.create_user(
            username=pending_user.username,
            email=pending_user.email,
            password=pending_user.password,
            ph_no=pending_user.ph_no,
            post=pending_user.post,
        )
        user.set_password(pending_user.password)
        user.save()

        # clean up pending user
        pending_user.delete()

        return Response(
            {"message": f"User {user.username} created successfully"}, status=201
        )

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(["POST"])
def login_view(request):
    serializer = LoginSerializer(data=request.data)

    if serializer.is_valid():
        user = serializer.validated_data["user"]

        if user:
            token, created = ExpiringToken.objects.get_or_create(user=user)

            if not created and token.is_expired():
                token.delete()
                token = ExpiringToken.objects.create(user=user)
                token.save()

        return Response(
            {
                "message": f"Welcome, {user.username}!",
                "token": token.key,
                "user_id": user.id,
                "expires_at": token.created + timedelta(minutes=1),
            },
            status=status.HTTP_200_OK,
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(["POST"])
def add_users_to_group(request):
    serializer = GroupCreateSerializers(data=request.data)

    if serializer.is_valid():
        group_name = serializer.validated_data["name"]
        user_ids = serializer.validated_data.get("users", [])

        # Check if the group exists
        group = CustomGroupModel.objects.filter(name=group_name).first()

        if not group:
            return Response(
                {"detail": "Group does not exist."}, status=status.HTTP_400_BAD_REQUEST
            )

        # Get users to add
        users = CustomUserModel.objects.filter(id__in=user_ids)

        added_users = 0

        for user in users:
            # Check if the user is already part of the group
            if group not in user.groups.all():
                user.groups.add(group)
                added_users += 1

        return Response(
            {"detail": f"{added_users} user(s) added to the group '{group_name}'."},
            status=status.HTTP_200_OK,
        )

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

#################################################################3
@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def user_list_create(request):

    if request.method == "GET":
        users = CustomUserModel.objects.all()
        serializer = UserSerializer(users, many=True)
        return Response(serializer.data)

    elif request.method == "POST":
        serializer = UserSerializer(data=request.data)

        if serializer.is_valid():
            serializer.save()

            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(["GET", "PUT", "DELETE"])
def user_detail(request, pk):
    try:
        user = CustomUserModel.objects.get(pk=pk)
    except CustomUserModel.DoesNotExist:
        return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

    if request.method == "GET":
        serializer = UserSerializer(user)
        return Response(serializer.data)
    elif request.method == "PUT":
        serializer = UserSerializer(user, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    elif request.method == "DELETE":
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

# Create GET API to fetch all user groups and their users
@api_view(["GET"])
def group_list_with_users(request):
    groups = CustomGroupModel.objects.all()
    serializer = GroupWithUsersSerializer(groups, many=True)
    return Response(serializer.data, status=status.HTTP_200_OK)

# Modify the GET API to filter users based on group ID
@api_view(["GET"])
def users_by_group(request, group_id=None):
    if group_id is not None:
        try:
            group = CustomGroupModel.objects.get(id=group_id)
            users = group.user_set.all()
            serializer = UserSerializer(users, many=True)
            return Response({"group": group.name, "users": serializer.data})
        except CustomGroupModel.DoesNotExist:
            return Response(
                {"error": "Group not found"}, status=status.HTTP_404_NOT_FOUND
            )

    # Return all groups with users if no group_id is passed
    groups_data = []
    for group in CustomGroupModel.objects.all():
        users = group.user_set.all()
        serializer = UserSerializer(users, many=True)
        groups_data.append({"group": group.name, "users": serializer.data})

    return Response(groups_data)

# fetch group data from name
@api_view(["GET"])
def get_groups(request):
    group_name = request.GET.get("name", None)
    groups = CustomGroupModel.objects.filter(name__icontains=group_name)

    if groups:

        users = set()
        for group in groups:
            for user in group.user_set.all():
                users.add(user)
        serializer = UserSerializer(users, many=True)
    else:
        groups = CustomGroupModel.objects.all()
        serializer = GroupSerializer(groups, many=True)

    return Response(serializer.data, status=status.HTTP_200_OK)

# class based Group CRUD operations
# Create and List Groups
class GroupListCreateView(generics.ListCreateAPIView):
    queryset = CustomGroupModel.objects.all()
    serializer_class = GroupSerializer

# Retrieve, Update, Delete a Group
class GroupDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = CustomGroupModel.objects.all()
    serializer_class = GroupSerializer


# Product and Category here
class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.filter(is_active=True)
    serializer_class = CategorySerializer  
    
    
# Custom filter set for Product
class ProductFilter(FilterSet):
    name = CharFilter(field_name="name", lookup_expr="icontains")

    class Meta:
        model = Product
        fields = ["name", "is_active"]

class ProductViewSet(viewsets.ModelViewSet):
    queryset = Product.objects.all().order_by("name")
    serializer_class = ProductSerializer

    filter_backends = [DjangoFilterBackend, OrderingFilter]
    filterset_class = ProductFilter

    ordering_fields = ["name", "-name"]  # Allow ascending and descending ordering
    ordering = ["name"]  # default ordering

class ProductCategoryViewSet(viewsets.ModelViewSet):
    queryset = ProductCategory.objects.all()
    serializer_class = ProductCategorySerializer

# order and payments
class PlaceOrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    # permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            # get id from order
            order_data = serializer.data
            order_id = order_data.get("id")
           
           # Send WebSocket message
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "products",
                {
                    "type": "order_placed",
                    "message": f"New order id {order_id} has been placed!"
                }
            )
            return Response({
                'message': 'Order placed successfully',
                'data': serializer.data
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

############payments###############
class PaymentViewSet(viewsets.ModelViewSet):
    queryset = Payment.objects.all()
    serializer_class = PaymentSerializer
    # permission_classes = [IsAuthenticated]

@require_GET
def order_list_view(request):
    page_number = request.GET.get('page',1)
    page_size = request.GET.get('page_size',2)
    orders = Order.objects.select_related('placed_by').prefetch_related('items__product').all().order_by('-created_at')

    paginator = Paginator(orders, page_size)
    page_obj = paginator.get_page(page_number)

    data = []
    for order in page_obj:
        order_data = {
            'order_id': order.id,
            'customer': order.customer,
            'total': float(order.total),
            'paid_amount': float(order.paid_amount),
            'status': order.status,
            'order_type': order.type,
            'shipping_address': order.shipping_address,
            'paid': order.paid,
            'created_at': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'placed_by': {
                'id': order.placed_by.id,
                'username': order.placed_by.username,
                'email': order.placed_by.email,
                "ph_no" : order.placed_by.ph_no,
                "address" : order.placed_by.address,
                "city" : order.placed_by.city,
                "country" : order.placed_by.country,
                "pincode" : order.placed_by.pincode,
                
            },
            'items': [],
            'payments' :[]
        }

        for item in order.items.all():
            order_data['items'].append({
                'product_id': item.product.id,
                'product_name': item.product.name,
                'price': float(item.price),
                'quantity': item.quantity
            })

        for payment in order.payment.all():
            order_data['payments'].append({
                'payment_id': payment.id,
                'payment_amount': float(payment.payment_amount),
                'payment_mode': payment.payment_mode,
                'reference_id': payment.reference_id,
                'created_at': payment.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        data.append(order_data)
    return JsonResponse({
        'page': page_obj.number,
        'total_pages': paginator.num_pages,
        'total_orders': paginator.count,
        'orders': data
    }, safe=False) # safe is use for except dict

# excel data from here 
from rest_framework.decorators import api_view
from django.http import HttpResponse
from io import BytesIO
import pandas as pd

@api_view(['GET'])
def product_data_download(request):
    try:
        product_categories = ProductCategory.objects.all()
        serializer = ProductCategorySerializer(product_categories, many=True)

        # Flatten the data
        flattened_data = [{
            'category_name': item['category']['name'].capitalize(),
            'category_description': item['category']['description'],
            'product_name': item['product']['name'].capitalize(),
            'product_description': item['product']['description'],
        } for item in serializer.data]

        # Create DataFrame
        df = pd.DataFrame(flattened_data)
        df_cleaned = df.sort_values(by=['category_name', 'product_name'])

        # Create Excel in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df_cleaned.to_excel(writer, sheet_name='products', startrow=1, header=False, index=False)

            workbook = writer.book
            worksheet = writer.sheets['products']
            header_format = workbook.add_format({'bold': True})

            for col_num, value in enumerate(df_cleaned.columns.values):
                worksheet.write(0, col_num, value, header_format)

        output.seek(0)

      # Define the file path
        media_dir = settings.MEDIA_ROOT
        file_path = os.path.join(media_dir, 'products.xlsx')

        # Create the directory if it doesn't exist
        os.makedirs(media_dir, exist_ok=True)


        # Save the workbook to the specified file
        with open(file_path, 'wb') as f:
            f.write(output.getvalue())

        # Return the link to the generated file
        file_url = request.build_absolute_uri(os.path.join(settings.MEDIA_URL, 'products.xlsx'))
        return Response({"download_link": file_url}, status=200)


    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)
        

@api_view(['POST'])
def product_data_upload(request):
    if request.FILES:
        try:
            wb = load_workbook(request.FILES['excel_file'])
            ws = wb['products']

            row = ws.max_row
            column = ws.max_column

            all_rows = list(ws.rows)
            wb.close()

            for row in all_rows[1:]:
                category_name = row[0].value
                # category_description = row[1].value
                product_name = row[2].value
                product_description = row[3].value

                product = Product.objects.create(name = product_name, description = product_description)
 
                try:
                    category_obj = Category.objects.get(name__iexact = category_name)
                    ProductCategory.objects.create(product=product, category=category_obj)
                except Category.DoesNotExist:
                    print(f"Category '{category_name}' not found")

        except Exception as e:
            return Response({'error':str(e)})
    return Response({'message':'file upload successfully'})

########store#######
class StoreViewSet(viewsets.ModelViewSet):
    queryset = Store.objects.all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return StoreCreateUpdateSerializer
        return StoreSerializer

    def perform_create(self, serializer):
        custom_user = CustomUserModel.objects.get(id=self.request.user.id)
        serializer.save(owner=custom_user)

    @action(detail=True, methods=['post'], permission_classes=[IsAuthenticated]) 
    def toggle_open(self, request, pk=None): #http://127.0.0.1:8000/stores/1/toggle_open/
        store = get_object_or_404(Store, pk=pk, owner=request.user)
        store.is_open = not store.is_open
        
        store.save()
        store.refresh_from_db()
        return Response({
            "status": "success",
            "store": store.store_name,
            "is_open": store.is_open,
            "is_currently_open": store.is_currently_open()
        })

class VisibleCategoriesView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, store_id):
        try:
            store = Store.objects.get(pk=store_id)
            if store.is_currently_open():
                categories = store.categories.filter(is_active=True)
                serializer = CategorySerializer(categories, many=True)
                return Response(serializer.data, status=status.HTTP_200_OK)
            else:
                return Response(
                    {"message": "Store is currently closed"},
                    status=status.HTTP_403_FORBIDDEN
                )
        except Store.DoesNotExist:
            return Response(
                {"error": "Store not found"},
                status=status.HTTP_404_NOT_FOUND
            )
